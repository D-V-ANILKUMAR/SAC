import os, json
from datetime import timedelta, datetime
from io import BytesIO
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
from werkzeug.utils import secure_filename
from dotenv import load_dotenv
import psycopg2
import psycopg2.extras
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import pandas as pd
from flask import send_file
from io import BytesIO
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from flask import request  # make sure this is imported

from flask import send_file, request, redirect, url_for, session


from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet


# Load .env
load_dotenv()
DATABASE_URL = os.environ.get('DATABASE_URL')
if not DATABASE_URL:
    raise RuntimeError("DATABASE_URL not set")

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET', 'secret123')
app.permanent_session_lifetime = timedelta(days=3650)
UPLOAD_FOLDER = 'static/uploads'
TAB_SWITCH_LIMIT = 3
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# --------------------
# DB Connection
# --------------------
def get_db_connection():
    return psycopg2.connect(DATABASE_URL, sslmode='require')

# --------------------
# DB Init / Schema
# --------------------
def init_db():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('''CREATE TABLE IF NOT EXISTS users (
                        id SERIAL PRIMARY KEY,
                        name TEXT,
                        email TEXT UNIQUE,
                        mobile TEXT,
                        password TEXT,
                        role TEXT
                    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS exams (
                        id SERIAL PRIMARY KEY,
                        title TEXT,
                        duration INTEGER,
                        created_by INTEGER,
                        attempts_allowed INTEGER DEFAULT 1
                    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS questions (
                        id SERIAL PRIMARY KEY,
                        exam_id INTEGER,
                        question TEXT,
                        image TEXT,
                        option1 TEXT,
                        option2 TEXT,
                        option3 TEXT,
                        option4 TEXT,
                        answer TEXT
                    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS submissions (
                        id SERIAL PRIMARY KEY,
                        exam_id INTEGER,
                        student_id INTEGER,
                        answers TEXT,
                        score INTEGER,
                        attempt_number INTEGER DEFAULT 1,
                        submitted_at TIMESTAMP
                    )''')
    # default accounts
    cur.execute('SELECT COUNT(*) FROM users')
    if cur.fetchone()[0] == 0:
        users = [
            ('Admin','admin@example.com','9999999999','admin123','admin'),
            ('Mediator1','mediator@example.com','8888888888','mediator123','mediator'),
            ('Student1','student@example.com','7777777777','student123','student')
        ]
        for u in users:
            cur.execute('INSERT INTO users (name,email,mobile,password,role) VALUES (%s,%s,%s,%s,%s)', u)
    conn.commit()
    cur.close()
    conn.close()

init_db()

# --------------------
# Helpers
# --------------------
def get_user(email):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('SELECT * FROM users WHERE email=%s', (email,))
    user = cur.fetchone()
    conn.close()
    return user

def get_all_users(role=None):
    conn = get_db_connection()
    cur = conn.cursor()
    if role:
        cur.execute('SELECT * FROM users WHERE role=%s', (role,))
    else:
        cur.execute('SELECT * FROM users')
    users = cur.fetchall()
    conn.close()
    return users

def get_exams():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT id, title, duration, created_by, attempts_allowed FROM exams')
    exams = [dict(r) for r in cur.fetchall()]
    conn.close()
    return exams

def get_exam(exam_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT id, title, duration, created_by, attempts_allowed FROM exams WHERE id=%s', (exam_id,))
    exam = dict(cur.fetchone()) if cur.rowcount > 0 else None
    cur.execute('SELECT id, exam_id, question, image, option1, option2, option3, option4, answer FROM questions WHERE exam_id=%s', (exam_id,))
    questions = [dict(r) for r in cur.fetchall()]
    conn.close()
    return exam, questions

# --------------------
# Routes
# --------------------
@app.route('/')
def index():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method=='POST':
        email = request.form['email']
        password = request.form['password']
        user = get_user(email)
        if user and user[4]==password:
            session.permanent = True
            session['user_id'] = user[0]
            session['role'] = user[5]
            session['name'] = user[1]
            session['tab_switch'] = 0
            if user[5]=='admin':
                return redirect(url_for('admin_home'))
            elif user[5]=='mediator':
                return redirect(url_for('mediator_home'))
            else:
                return redirect(url_for('student_home'))
        else:
            flash('Invalid credentials')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# --------------------
# Admin Routes
# --------------------
@app.route('/admin/home')
def admin_home():
    if session.get('role')!='admin': 
        return redirect(url_for('login'))
    return render_template('admin_home.html', name=session['name'])


# --------------------
# Admin Routes (Add these to your existing admin routes section)
# --------------------

@app.route('/admin/create_exam', methods=['GET','POST'])
def create_exam():
    if session.get('role') != 'admin':
        return redirect(url_for('login'))
        
    if request.method=='POST':
        title = request.form['title']
        duration = int(request.form['duration'])
        attempts_allowed = int(request.form.get('attempts_allowed', 1))
        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute(
                "INSERT INTO exams (title,duration,created_by,attempts_allowed) VALUES (%s,%s,%s,%s) RETURNING id",
                (title,duration,session['user_id'],attempts_allowed)
            )
            exam_id = cur.fetchone()[0]
            questions_count = int(request.form['qcount'])
            for i in range(1, questions_count+1):
                question = request.form[f'question{i}']
                opt1 = request.form[f'opt1_{i}']
                opt2 = request.form[f'opt2_{i}']
                opt3 = request.form[f'opt3_{i}']
                opt4 = request.form[f'opt4_{i}']
                answer = request.form[f'answer{i}']
                img_file = request.files.get(f'image{i}')
                filename = None
                if img_file and img_file.filename != '':
                    filename = secure_filename(img_file.filename)
                    img_file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                cur.execute(
                    "INSERT INTO questions (exam_id,question,image,option1,option2,option3,option4,answer) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                    (exam_id,question,filename,opt1,opt2,opt3,opt4,answer)
                )
            conn.commit()
            flash("Exam created successfully!")
        except Exception as e:
            conn.rollback()
            flash(f"Error: {e}")
        finally:
            cur.close()
            conn.close()
        return redirect(url_for('admin_home'))
    return render_template('create_exam_admin.html')

@app.route('/admin/submissions')
def admin_submissions():
    if session.get('role') != 'admin':
        return redirect(url_for('login'))
        
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('''
        SELECT s.*, u.name as student_name, e.title as exam_title 
        FROM submissions s 
        JOIN users u ON s.student_id = u.id 
        JOIN exams e ON s.exam_id = e.id 
        ORDER BY s.submitted_at DESC
    ''')
    submissions = cur.fetchall()
    conn.close()
    return render_template('admin_submissions.html', submissions=submissions)

@app.route('/admin/create_mediator', methods=['GET','POST'])
def create_mediator():
    if session.get('role') != 'admin':
        return redirect(url_for('login'))

    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        mobile = request.form['mobile']
        password = request.form['password']

        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute(
                "INSERT INTO users (name,email,mobile,password,role) VALUES (%s,%s,%s,%s,'mediator')",
                (name,email,mobile,password)
            )
            conn.commit()
            flash("Mediator created successfully!")
        except Exception as e:
            conn.rollback()
            flash(f"Error creating mediator: {e}")
        finally:
            cur.close()
            conn.close()
        return redirect(url_for('admin_home'))

    return render_template('create_mediator.html')

@app.route('/admin/create_student', methods=['GET','POST'])
def create_student_admin():
    if session.get('role') != 'admin':
        return redirect(url_for('login'))

    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        mobile = request.form['mobile']
        password = request.form['password']

        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute(
                "INSERT INTO users (name,email,mobile,password,role) VALUES (%s,%s,%s,%s,'student')",
                (name,email,mobile,password)
            )
            conn.commit()
            flash("Student created successfully!")
        except Exception as e:
            conn.rollback()
            flash(f"Error creating student: {e}")
        finally:
            cur.close()
            conn.close()
        return redirect(url_for('admin_home'))

    return render_template('create_student_admin.html')

# Admin account details
@app.route('/admin/account_details')
def account_details_admin():
    if session.get('role') != 'admin':
        return redirect(url_for('login'))

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT id,name,email,mobile,password,role FROM users")
    users = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('account_details_admin.html', users=users)

# Delete user
@app.route('/admin/delete_user', methods=['POST'])
def delete_user():
    if session.get('role') != 'admin':
        return redirect(url_for('login'))

    user_id = request.form.get('user_id')
    if str(user_id) == str(session.get('user_id')):
        flash("You cannot delete your own account.")
        return redirect(url_for('account_details_admin'))

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM users WHERE id = %s", (user_id,))
    conn.commit()
    cur.close()
    conn.close()
    flash("User deleted successfully.")
    return redirect(url_for('account_details_admin'))
@app.route('/admin/manage_exams')
def manage_exams_admin():
    if session.get('role') != 'admin' :
        return redirect(url_for('login'))
        
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT id,title,duration,created_by,attempts_allowed FROM exams")
    exams = cur.fetchall()
    conn.close()
    return render_template('manage_exams_admin.html', exams=exams)

@app.route('/admin/leaderboard')
def admin_leaderboard():
    if session.get('role') != 'admin':
        return redirect(url_for('login'))

    exam_id = request.args.get('exam_id', type=int)  # get exam filter from query params

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Get list of all exams for dropdown
    cur.execute("SELECT id, title FROM exams ORDER BY title;")
    exams = cur.fetchall()

    # Fetch leaderboard with new sorting: score DESC, time_taken ASC, submitted_at ASC
    if exam_id:
        cur.execute('''
            SELECT s.id, u.name as student_name, e.title as exam_title, s.score, 
                   s.attempt_number, s.submitted_at, s.time_taken
            FROM submissions s
            JOIN users u ON s.student_id = u.id
            JOIN exams e ON s.exam_id = e.id
            WHERE e.id = %s
            ORDER BY s.score DESC, 
                     COALESCE(s.time_taken, 999999) ASC, 
                     s.submitted_at ASC
        ''', (exam_id,))
    else:
        cur.execute('''
            SELECT s.id, u.name as student_name, e.title as exam_title, s.score, 
                   s.attempt_number, s.submitted_at, s.time_taken
            FROM submissions s
            JOIN users u ON s.student_id = u.id
            JOIN exams e ON s.exam_id = e.id
            ORDER BY s.score DESC, 
                     COALESCE(s.time_taken, 999999) ASC, 
                     s.submitted_at ASC
        ''')
    
    leaderboard = cur.fetchall()
    conn.close()

    return render_template(
        'admin_leaderboard.html',
        leaderboard=leaderboard,
        exams=exams,
        selected_exam=exam_id
    )
from io import BytesIO
from flask import send_file, request, redirect, url_for, session
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

@app.route('/admin/leaderboard/download/pdf')
def download_leaderboard_pdf():
    if session.get('role') != 'admin':
        return redirect(url_for('login'))
    
    exam_id = request.args.get('exam_id', type=int)

    conn = get_db_connection()
    cur = conn.cursor()

    # Fixed sorting: score DESC, time_taken ASC, submitted_at ASC
    if exam_id:
        cur.execute('''
            SELECT u.name, e.title, s.score, s.attempt_number, s.submitted_at, s.time_taken
            FROM submissions s
            JOIN users u ON s.student_id = u.id
            JOIN exams e ON s.exam_id = e.id
            WHERE e.id = %s
            ORDER BY s.score DESC, COALESCE(s.time_taken, 999999) ASC, s.submitted_at ASC
        ''', (exam_id,))
    else:
        cur.execute('''
            SELECT u.name, e.title, s.score, s.attempt_number, s.submitted_at, s.time_taken
            FROM submissions s
            JOIN users u ON s.student_id = u.id
            JOIN exams e ON s.exam_id = e.id
            ORDER BY s.score DESC, COALESCE(s.time_taken, 999999) ASC, s.submitted_at ASC
        ''')
    
    data = cur.fetchall()
    conn.close()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    elements.append(Paragraph("Leaderboard Report", styles['Title']))
    elements.append(Spacer(1, 12))

    # Table header + data
    table_data = [["S.No", "Student Name", "Exam Title", "Score", "Attempt", "Submitted At", "Time Taken"]]
    for idx, row in enumerate(data, start=1):
        submitted_at_str = row[4].strftime('%Y-%m-%d %H:%M') if row[4] else "Not Submitted"
        time_taken_str = f"{row[5] // 60}m {row[5] % 60}s" if row[5] else "-"
        table_data.append([idx, row[0], row[1], row[2], row[3], submitted_at_str, time_taken_str])

    # Table styling
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#4FACFE")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 11),
        ('BOTTOMPADDING', (0,0), (-1,0), 10),
        ('BACKGROUND', (0,1), (-1,-1), colors.whitesmoke),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="leaderboard.pdf",
        mimetype="application/pdf"
    )
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from flask import send_file, request, redirect, url_for, session

@app.route('/admin/leaderboard/download/excel')
def download_leaderboard_excel():
    if session.get('role') != 'admin':
        return redirect(url_for('login'))

    exam_id = request.args.get('exam_id', type=int)

    conn = get_db_connection()
    cur = conn.cursor()

    # Fixed sorting: score DESC, time_taken ASC, submitted_at ASC
    if exam_id:
        cur.execute('''
            SELECT u.name, e.title, s.score, s.attempt_number, s.submitted_at, s.time_taken
            FROM submissions s
            JOIN users u ON s.student_id = u.id
            JOIN exams e ON s.exam_id = e.id
            WHERE e.id = %s
            ORDER BY s.score DESC, COALESCE(s.time_taken, 999999) ASC, s.submitted_at ASC
        ''', (exam_id,))
    else:
        cur.execute('''
            SELECT u.name, e.title, s.score, s.attempt_number, s.submitted_at, s.time_taken
            FROM submissions s
            JOIN users u ON s.student_id = u.id
            JOIN exams e ON s.exam_id = e.id
            ORDER BY s.score DESC, COALESCE(s.time_taken, 999999) ASC, s.submitted_at ASC
        ''')
    
    data = cur.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leaderboard"

    headers = ["S.No", "Student Name", "Exam Title", "Score", "Attempt", "Submitted At", "Time Taken"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4FACFE", end_color="4FACFE", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    # Header row
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border

    # Data rows
    for idx, row in enumerate(data, start=1):
        submitted_at_str = row[4].strftime('%Y-%m-%d %H:%M') if row[4] else "Not Submitted"
        time_taken_str = f"{row[5] // 60}m {row[5] % 60}s" if row[5] else "-"
        ws.append([idx, row[0], row[1], row[2], row[3], submitted_at_str, time_taken_str])

    # Apply alignment & borders to all data cells
    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row_cells:
            cell.alignment = alignment
            cell.border = thin_border

    # Adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="leaderboard.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
@app.route('/admin/bulk_upload', methods=['GET', 'POST'])
def admin_bulk_upload():
    if session.get('role') != 'admin':
        return redirect(url_for('login'))

    if request.method == 'GET':
        return render_template('bulk_upload.html', name=session.get('name', 'Admin'))

    uploaded = request.files.get('file')
    if not uploaded or uploaded.filename == '':
        flash("No file selected", "danger")
        return redirect(url_for('admin_bulk_upload'))

    filename = uploaded.filename.lower()
    if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
        flash("Unsupported file type. Use CSV/XLS/XLSX.", "danger")
        return redirect(url_for('admin_bulk_upload'))

    try:
        if filename.endswith('.csv'):
            df = pd.read_csv(uploaded, dtype=str)
        else:
            df = pd.read_excel(uploaded, dtype=str)

        df = df.replace({pd.NA: None, 'nan': None, 'NaN': None})
        df.columns = [c.strip().lower() for c in df.columns]
    except Exception as e:
        flash(f"Error reading file: {str(e)}", "danger")
        return redirect(url_for('admin_bulk_upload'))

    required_cols = ['email', 'password']
    for col in required_cols:
        if col not in df.columns:
            flash(f"Missing required column: {col}", "danger")
            return redirect(url_for('admin_bulk_upload'))

    created, failed, skipped = 0, 0, 0
    results = []
    conn = get_db_connection()
    cur = conn.cursor()

    for idx, row in df.iterrows():
        try:
            email = str(row.get('email') or '').strip()
            password = str(row.get('password') or '').strip()
            name = str(row.get('name') or '').strip()
            mobile = str(row.get('mobile') or '').strip()
            role = (row.get('role') or 'student').strip().lower()

            if not email or not password:
                skipped += 1
                continue
            if '@' not in email:
                failed += 1
                continue
            if role not in ['student', 'mediator', 'admin']:
                role = 'student'

            cur.execute(
                "INSERT INTO users (name, email, mobile, password, role) VALUES (%s,%s,%s,%s,%s)",
                (name or None, email, mobile or None, password, role)
            )
            conn.commit()
            created += 1
        except psycopg2.IntegrityError:
            conn.rollback()
            failed += 1
        except Exception as e:
            conn.rollback()
            failed += 1

    cur.close()
    conn.close()

    flash(f"Upload finished: {created} created, {skipped} skipped, {failed} failed",
          "success" if failed == 0 else "warning")
    return redirect(url_for('admin_bulk_upload'))


@app.route('/admin/manage_exams/edit/<int:exam_id>', methods=['GET', 'POST'])
def edit_exam(exam_id):
    if session.get('role') != 'admin':
        return redirect(url_for('login'))

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Fetch exam info
    cur.execute("SELECT id, title, duration, attempts_allowed FROM exams WHERE id = %s", (exam_id,))
    exam = cur.fetchone()
    if not exam:
        conn.close()
        flash("Exam not found", "danger")
        return redirect(url_for("manage_exams_admin"))

    # Fetch questions
    cur.execute("""
        SELECT id, question, option1, option2, option3, option4, answer
        FROM questions
        WHERE exam_id = %s
    """, (exam_id,))
    questions = cur.fetchall()
    print("DEBUG: Questions fetched:", questions)  # Debugging

    if request.method == "POST":
        # Update exam info
        cur.execute("""
            UPDATE exams
            SET title=%s, duration=%s, attempts_allowed=%s
            WHERE id=%s
        """, (request.form['title'], request.form['duration'], request.form['attempts_allowed'], exam_id))

        # Update questions
        for q in questions:
            q_id = q['id']
            cur.execute("""
                UPDATE questions
                SET question=%s, option1=%s, option2=%s, option3=%s, option4=%s, answer=%s
                WHERE id=%s
            """, (
                request.form.get(f"question_{q_id}"),
                request.form.get(f"option1_{q_id}"),
                request.form.get(f"option2_{q_id}"),
                request.form.get(f"option3_{q_id}"),
                request.form.get(f"option4_{q_id}"),
                request.form.get(f"answer_{q_id}"),
                q_id
            ))

        conn.commit()
        conn.close()
        flash("Exam and questions updated successfully!", "success")
        return redirect(url_for("manage_exams_admin"))

    conn.close()
    return render_template("edit_exam.html", exam=exam, questions=questions)

# ---------- DELETE ----------
@app.route('/admin/manage_exams/delete/<int:exam_id>')
def delete_exam(exam_id):
    if session.get('role') != 'admin':
        return redirect(url_for('login'))

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM exams WHERE id = %s", (exam_id,))
    conn.commit()
    conn.close()
    flash("Exam deleted successfully!", "success")
    return redirect(url_for("manage_exams_admin"))




# --------------------
# Mediator Routes
# --------------------
@app.route('/mediator/home')
def mediator_home():
    if session.get('role')!='mediator': 
        return redirect(url_for('login'))
    return render_template('mediator_home.html', name=session['name']) 

@app.route('/mediator/manage_exams')
def manage_exams_mediator():
    if 'role' not in session or session['role'] != 'mediator':
        return redirect(url_for('login'))

    mediator_id = session.get('user_id')  # Make sure this is set in session

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Fetch exams created by this mediator
    cur.execute("""
        SELECT id, title, duration, created_by, attempts_allowed
        FROM exams
        WHERE created_by = %s
    """, (mediator_id,))
    exams = cur.fetchall()
    conn.close()

    return render_template('manage_exams_mediator.html', exams=exams)



@app.route('/mediator/create_student', methods=['GET','POST'])
def create_student_mediator():
    if session.get('role') != 'mediator':
        return redirect(url_for('login'))

    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        mobile = request.form['mobile']
        password = request.form['password']

        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute(
                "INSERT INTO users (name,email,mobile,password,role) VALUES (%s,%s,%s,%s,'student')",
                (name,email,mobile,password)
            )
            conn.commit()
            flash("Student created successfully!")
        except Exception as e:
            conn.rollback()
            flash(f"Error creating student: {e}")
        finally:
            cur.close()
            conn.close()
        return redirect(url_for('mediator_home'))

    return render_template('create_student.html')

@app.route('/mediator/account_details')
def account_details_mediator():
    if session.get('role') != 'mediator':
        return redirect(url_for('login'))

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT id,name,email,mobile,password FROM users WHERE role='student'")
    students = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('account_details_mediator.html', students=students)

@app.route('/mediator/create_exam', methods=['GET','POST'])
def create_exam_mediator():
    if session.get('role') != 'mediator':
        return redirect(url_for('login'))
        
    if request.method=='POST':
        title = request.form['title']
        duration = int(request.form['duration'])
        attempts_allowed = int(request.form.get('attempts_allowed', 1))
        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute(
                "INSERT INTO exams (title,duration,created_by,attempts_allowed) VALUES (%s,%s,%s,%s) RETURNING id",
                (title,duration,session['user_id'],attempts_allowed)
            )
            exam_id = cur.fetchone()[0]
            questions_count = int(request.form['qcount'])
            for i in range(1, questions_count+1):
                question = request.form[f'question{i}']
                opt1 = request.form[f'opt1_{i}']
                opt2 = request.form[f'opt2_{i}']
                opt3 = request.form[f'opt3_{i}']
                opt4 = request.form[f'opt4_{i}']
                answer = request.form[f'answer{i}']
                img_file = request.files.get(f'image{i}')
                filename = None
                if img_file and img_file.filename != '':
                    filename = secure_filename(img_file.filename)
                    img_file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                cur.execute(
                    "INSERT INTO questions (exam_id,question,image,option1,option2,option3,option4,answer) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                    (exam_id,question,filename,opt1,opt2,opt3,opt4,answer)
                )
            conn.commit()
            flash("Exam created successfully!")
        except Exception as e:
            conn.rollback()
            flash(f"Error: {e}")
        finally:
            cur.close()
            conn.close()
        return redirect(url_for('mediator_home'))
    return render_template('create_exam.html')




# --------------------
# Student Routes
# --------------------
@app.route('/student/home')
def student_home():
    if session.get('role')!='student': 
        return redirect(url_for('login'))
        
    exams = get_exams()
    uid = session.get('user_id')
    conn = get_db_connection()
    cur = conn.cursor()
    exam_infos = []
    for ex in exams:
        ex_id = ex['id']
        attempts_allowed = ex.get('attempts_allowed',1)
        cur.execute('SELECT COUNT(*) FROM submissions WHERE exam_id=%s AND student_id=%s', (ex_id, uid))
        prev_attempts = cur.fetchone()[0]
        remaining = max(attempts_allowed - prev_attempts, 0)
        exam_infos.append({
            'id': ex_id,
            'title': ex['title'],
            'duration': ex['duration'],
            'created_by': ex.get('created_by'),
            'attempts_allowed': attempts_allowed,
            'prev_attempts': prev_attempts,
            'remaining': remaining,
        })
    cur.close()
    conn.close()
    return render_template('student_home.html', exam_infos=exam_infos, name=session.get('name'))

from datetime import datetime
import json

@app.route('/student/take_exam/<int:exam_id>', methods=['GET','POST'])
def take_exam(exam_id):
    if session.get('role') != 'student': 
        return redirect(url_for('login'))
        
    exam, questions = get_exam(exam_id)
    if not exam:
        flash('Exam not found')
        return redirect(url_for('student_home'))
        
    uid = session.get('user_id')

    # Check attempts
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('SELECT attempts_allowed FROM exams WHERE id=%s', (exam_id,))
    result = cur.fetchone()
    attempts_allowed = result[0] if result else 1
    cur.execute('SELECT COUNT(*) FROM submissions WHERE exam_id=%s AND student_id=%s', (exam_id, uid))
    prev_attempts = cur.fetchone()[0]
    
    if prev_attempts >= attempts_allowed:
        flash(f'You have already attempted this exam ({attempts_allowed} allowed).')
        cur.close()
        conn.close()
        return redirect(url_for('student_home'))

    if request.method == 'POST':
        answers = {}
        for q in questions:
            val = request.form.get(f"q{q['id']}")
            answers[str(q['id'])] = val

        # calculate score
        score = sum(1 for q in questions if answers.get(str(q['id'])) == q['answer'])
        attempt_number = prev_attempts + 1
        submitted_at = datetime.now()

        # ⏱ fetch time_taken from hidden field (set by JS timer)
        time_taken = request.form.get('time_taken', type=int)

        cur.execute('''
            INSERT INTO submissions 
            (exam_id, student_id, answers, score, attempt_number, submitted_at, time_taken)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        ''', (exam_id, uid, json.dumps(answers), score, attempt_number, submitted_at, time_taken))

        conn.commit()
        cur.close()
        conn.close()
        flash('Exam submitted successfully!')
        return redirect(url_for('student_home'))

    cur.close()
    conn.close()
    return render_template(
        'take_exam.html',
        exam=exam,
        questions=questions,
        tab_limit=TAB_SWITCH_LIMIT,
        exam_duration=exam['duration']
    )


# --------------------
# Run
# --------------------
if __name__=='__main__':
    app.run(debug=True)